from logging import error
from sqlite3.dbapi2 import Error
from bs4 import BeautifulSoup
import requests
from datetime import datetime, timedelta
import sqlite3
import numpy_financial as npf
import numpy as np
import pandas as pd
import openpyxl
import requests

def generar_panel():

    comision = 0.51/100

    #Crear Excel

    wb = openpyxl.Workbook()
    hoja = wb.active
    hoja.cell(row=1, column=1, value="Especie")
    hoja.cell(row=1, column=2, value="Emisor")
    hoja.cell(row=1, column=3, value="Cotizacion")
    hoja.cell(row=1, column=4, value="Duration")
    hoja.cell(row=1, column=5, value="TIR")
    hoja.cell(row=1, column=6, value="Volumen")
    hoja.cell(row=1, column=7, value="Plazo")
    hoja.cell(row=1, column=8, value="Ley")
    hoja2 = wb.create_sheet("Hoja2", 0)
    hoja2.cell(row=1, column=1, value="Especie")

    fila = 2
    fila_2=2


    # Acceder a los datos del mercado

    try:
        r = requests.get('https://bolsar.info/Obligaciones_Negociables.php')

        soup = BeautifulSoup(r.text,features="html.parser")

        table = soup.find_all("table", {"id": "lideres1"})

        if not table:
            print("No anda la pagina")
        else:
            text = []
            for row in table[0].find_all("tr"):
                for cell in row.find_all("td"):
                    text.append(cell.get_text().strip())



        # ARMADO DE PANEL
        panel = {}

        for x in range(0,len(text),16):

            if text[x+15] == "":
                hora="00:00:00"
            else:
                hora=text[x+15]
            
            panel[text[x]] = {
                            'Plazo'           :   text[x+1],
                            'Cantidad Nominal':   int(text[x+2].replace(".","")),
                            'Compra'          :   float((text[x+3].replace(".","")).replace(",",".")),
                            'Venta'           :   float((text[x+4].replace(".","")).replace(",",".")),
                            'Cantidad Nominal':   int(text[x+5].replace(".","")),
                            'Ultimo'          :   float(((text[x+6].replace(".","")).replace(",",".")).replace("-","0")),
                            'Variacion'       :   float(((text[x+7].replace(".","").replace("%","")).replace(",","."))),
                            'Apertura'        :   float((text[x+8].replace(".","")).replace(",",".")),
                            'Maxima'          :   float(((text[x+9].replace(".","")).replace(",",".")).replace("-","0")),
                            'Minimo'          :   float(((text[x+10].replace(".","")).replace(",",".")).replace("-","0")),
                            'Cierre Anterior' :   float(((text[x+11].replace(".","")).replace(",",".")).replace("-","0")),
                            'Volumen'         :   int(text[x+12].replace(".","")),
                            'Monto'           :   float((text[x+13].replace(".","")).replace(",",".")),
                            'Operaciones'     :   int((text[x+14].replace(".","")).replace("","0")),
                            'Hora'            :   datetime.strptime(hora,'%H:%M:%S')}
            
            
            hoja2.cell(row=fila_2, column=1, value=text[x])
            fila_2+=1
                        
        conexion = sqlite3.connect('base_datos.db')
        cursor = conexion.cursor()
        cursor.execute('SELECT * FROM especies')
        rows = cursor.fetchall()


        for x in rows:
            
            if x[0] in panel:

                if bool(x[9]):

                    #Traigo de la BBDD los datos de cada ON del Panel
                    cursor_2 = conexion.cursor()
                    cursor_2.execute('SELECT * FROM flujo_fondos WHERE especie = "{}"'.format(x[0]))
                    rows_2 = cursor_2.fetchall()
                    

                    #Fechas
                    vencimiento = datetime.strptime(x[6],'%Y/%m/%d')
                    fecha_hoy = datetime(datetime.today().year,datetime.today().month,datetime.today().day)
                    duration = (vencimiento-fecha_hoy).days+1
                    tir_pagos = [panel[x[0]]['Ultimo']*(-1)*(1+ comision * 0.21)]
                    flujo_fondos = []


                    #Creo el FF desde la BBDD
                    for i in rows_2:
                        
                        flujo_fondos.append({datetime.strptime(i[1],'%Y/%m/%d'):i[2]})


                    #Armo un FF diario
                    for f in range(1, duration):

                        tir_pagos.append(0)

                        for i in range(len(flujo_fondos)):

                            if flujo_fondos[i].get(fecha_hoy+timedelta(days=f), False):
                                tir_pagos.pop()
                                tir_pagos.append(flujo_fondos[i].get(
                                    fecha_hoy+timedelta(days=f)))

                    #Calculo la TIR
                    tir = round(((1+npf.irr(tir_pagos))**365-1)*100,2)
                    duration = duration-1
                    especie = x[0]
                    emisor = x[1]
                    volumen = panel[x[0]]['Volumen']
                    cotizacion=panel[x[0]]['Ultimo']
                    plazo = panel[x[0]]['Plazo']
                    ley = x[11]


                    hoja.cell(row=fila, column=1, value=especie)
                    hoja.cell(row=fila, column=2, value=emisor)
                    hoja.cell(row=fila, column=3, value=cotizacion)
                    hoja.cell(row=fila, column=4, value=duration)
                    hoja.cell(row=fila, column=5, value=tir)
                    hoja.cell(row=fila, column=6, value=volumen)
                    hoja.cell(row=fila, column=7, value=plazo)
                    hoja.cell(row=fila, column=8, value=ley)
                    
                    fila += 1
            else:
                if bool(x[9]):

                    vencimiento = datetime.strptime(x[6],'%Y/%m/%d')
                    fecha_hoy = datetime(datetime.today().year,datetime.today().month,datetime.today().day)
                    duration = (vencimiento-fecha_hoy).days+1

                    tir = 0
                    duration = duration-1
                    especie = x[0]
                    emisor = x[1]
                    volumen = 0
                    cotizacion= 0
                    plazo = ""

                    hoja.cell(row=fila, column=1, value=especie)
                    hoja.cell(row=fila, column=2, value=emisor)
                    hoja.cell(row=fila, column=3, value=cotizacion)
                    hoja.cell(row=fila, column=4, value=duration)
                    hoja.cell(row=fila, column=5, value=tir)
                    hoja.cell(row=fila, column=6, value=volumen)
                    hoja.cell(row=fila, column=7, value=plazo)
                    hoja.cell(row=fila, column=8, value=ley)

                    fila += 1

                    texto = "Especie: {} TIR: {}".format(especie,tir)



        wb.save('panel_on.xlsx')

    #    r = requests.post('https://api.telegram.org/bot1862802635:AAFNYGw065RZFjINXEGVfEk6L1RIdNtEprE/sendMessage',
    #    data={'chat_id': 1595176733, 'text': "Excel"})
    #    r = requests.post('https://api.telegram.org/bot1862802635:AAFNYGw065RZFjINXEGVfEk6L1RIdNtEprE/sendMessage',
    #    data={'chat_id': 1301321058, 'text': "Excel"})    
    #
    #    r = requests.post('https://api.telegram.org/bot1862802635:AAFNYGw065RZFjINXEGVfEk6L1RIdNtEprE/sendDocument',
    #              files={'document': ('panel_on.xlsx', open('panel_on.xlsx', 'rb'))},
    #              data={'chat_id': 1595176733, 'caption': "Excel" })
    #    r = requests.post('https://api.telegram.org/bot1862802635:AAFNYGw065RZFjINXEGVfEk6L1RIdNtEprE/sendDocument',
    #              files={'document': ('panel_on.xlsx', open('panel_on.xlsx', 'rb'))},
    #              data={'chat_id': 1301321058, 'caption': "Excel" })

    #    data = json.loads(r.text)
    #    print(data['ok'])
        return "El archivo se genero correctamente."
    except Error as e:
        print(e)
        return("Reintente mas tarde")

if __name__ == "__main__":
    generar_panel()